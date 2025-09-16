# app.py — 산업용 공급량 예측(추세분석)
# • 데이터: 연도별 엑셀 여러 개 업로드(또는 Repo의 산업용_*.xlsx 자동 로딩)
# • 전처리: '상품명' == '산업용' (정확일치)만 사용, 집계는 '업종' 기준으로 '판매량' 합계
# • 월별 추정: 2025년 미확정 월(9–12월)을 같은 달 연도추세로 예측해 2025 연간(추정) 산출
# • 좌측: 학습 연도(멀티, 2020~선택), 예측 구간(시작연~종료연)
# • 예측: OLS / CAGR / Holt / SES — 다년 예측(연간)
# • 결과 유지: session_state 저장
# • 그래프: 총합(실적+예측포인트), Top-10 막대(연도 선택), Top-10 실적추이(예측연도 연장)
# • 다운로드: 전체표 + 방법별 시트(Top-20 막대, 연도별 총합 라인)

from pathlib import Path
from io import BytesIO
import re
import numpy as np
import pandas as pd
import streamlit as st
import altair as alt

# statsmodels (Holt, SES)
try:
    from statsmodels.tsa.holtwinters import Holt, SimpleExpSmoothing
except Exception:
    Holt = None
    SimpleExpSmoothing = None

# openpyxl (엑셀 차트)
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, LineChart

# ───────────────────────── 기본 UI ─────────────────────────
st.set_page_config(page_title="산업용 공급량 예측(추세분석)", layout="wide")
st.title("🏭📈 산업용 공급량 예측(추세분석)")
st.caption("연도별 엑셀 업로드(여러 개) 또는 Repo 일괄 로딩 → ‘산업용’만 필터 → 월별 보정으로 2025 연간(추정) 완성 → 4가지 추세 예측")

st.markdown(
    """
### 📘 예측 방법 설명
- **선형추세(OLS)** — `y_t = a + b t`, `ŷ_{T+h} = a + b (T+h)`
- **CAGR(복리성장)** — `g = (y_T / y_0)^{1/n} - 1`, `ŷ_{T+h} = y_T (1+g)^h`
- **Holt(지수평활·추세형)** — `ŷ_{T+h} = l_T + h b_T` (계절성 제외, 최근 추세를 더 반영)
- **지수평활(SES)** — `ŷ_{T+h} = l_T` (추세·계절성 없음)
"""
)

# ───────────────────────── 사이드바 ─────────────────────────
with st.sidebar:
    st.header("📥 ① 데이터 불러오기")
    ups = st.file_uploader("연도별 엑셀(.xlsx) 여러 개 업로드", type=["xlsx"], accept_multiple_files=True)

    st.caption("또는, Repo에 있는 **산업용_*.xlsx** 파일을 자동으로 읽을 수 있어.")
    repo_files = sorted([p for p in Path(".").glob("산업용_*.xlsx")])
    use_repo = st.checkbox(f"Repo 파일 자동 읽기 ({len(repo_files)}개 감지됨)", value=bool(repo_files))
    if repo_files:
        st.write("읽을 대상:", "\n\n".join([f"- {p.name}" for p in repo_files]))

    st.divider()
    st.header("🧪 ② 예측 방법")
    METHOD_CHOICES = ["선형추세(OLS)", "CAGR(복리성장)", "Holt(지수평활)", "지수평활(SES)"]
    methods = st.multiselect("방법 선택(정렬 기준은 첫 번째)", METHOD_CHOICES, default=METHOD_CHOICES)

# ───────────────────────── 유틸 ─────────────────────────
def _clean_col(s: str) -> str:
    return re.sub(r"\s+", "", str(s)).strip()

def _extract_year_from_filename(name: str) -> int | None:
    m = re.search(r"(19|20)(\d{2})(\d{2})?$", name.replace(".xlsx",""))
    if m: return int(m.group(1)+m.group(2))
    return None

def _parse_year_month(s: pd.Series) -> tuple[pd.Series, pd.Series]:
    """판매년월을 YYYY, M으로 파싱."""
    if s is None:
        return pd.Series([], dtype="Int64"), pd.Series([], dtype="Int64")
    x = s.astype(str).str.strip()
    dt = pd.to_datetime(x, errors="coerce", infer_datetime_format=True)
    if dt.isna().all():
        dt = pd.to_datetime(x, format="%b-%y", errors="coerce")  # Jan-25
    yy = dt.dt.year.astype("Int64")
    mm = dt.dt.month.astype("Int64")
    # 두 자리 숫자만 → 연도 보정, 월은 모를 수 있으니 남김
    mask_yy2 = yy.isna() & x.str.fullmatch(r"\d{2}")
    yy.loc[mask_yy2] = 2000 + x.loc[mask_yy2].astype(int)
    mask_yyyy = yy.isna() & x.str.fullmatch(r"(19|20)\d{2}")
    yy.loc[mask_yyyy] = x.loc[mask_yyyy].astype(int)
    return yy, mm

def _coerce_num(s):
    return pd.to_numeric(s, errors="coerce")

def _ols(x_years, y_vals, targets):
    coef = np.polyfit(x_years, y_vals, 1)
    fitted = np.polyval(coef, x_years)
    preds = [float(np.polyval(coef, t)) for t in targets]
    return preds, fitted

def _cagr(x_years, y_vals, targets):
    y0, yT = float(y_vals[0]), float(y_vals[-1]); n = int(x_years[-1] - x_years[0])
    if y0 <= 0 or yT <= 0 or n <= 0: return _ols(x_years, y_vals, targets)
    g = (yT / y0) ** (1.0 / n) - 1.0
    last = x_years[-1]
    preds = [float(yT * (1.0 + g) ** (t - last)) for t in targets]
    return preds, np.array(y_vals, dtype=float)

def _holt_level(y_vals, x_years, targets, damped=True):
    """연도축 단변량에 Holt 적용(월별 예측용, 필요 시 OLS로 fallback)."""
    if Holt is None or len(y_vals) < 2:
        return _ols(x_years, y_vals, targets)
    fit = Holt(np.asarray(y_vals), exponential=False, damped_trend=damped,
               initialization_method="estimated").fit(optimized=True)
    last = x_years[-1]
    steps = [t - last for t in targets]
    fc = fit.forecast(max(steps))
    preds = [float(fc[h-1]) for h in steps]
    return preds, np.array(fit.fittedvalues, dtype=float)

def _holt_annual(y_vals, last_train_year, targets, damped=False):
    if Holt is None or len(y_vals) < 2 or any(t <= last_train_year for t in targets):
        return _ols(list(range(last_train_year - len(y_vals) + 1, last_train_year + 1)), y_vals, targets)
    fit = Holt(np.asarray(y_vals), exponential=False, damped_trend=damped,
               initialization_method="estimated").fit(optimized=True)
    max_h = max(t - last_train_year for t in targets)
    fc = fit.forecast(max_h)
    preds = [float(fc[h - 1]) for h in [t - last_train_year for t in targets]]
    return preds, np.array(fit.fittedvalues, dtype=float)

def _ses(y_vals, last_train_year, targets):
    if SimpleExpSmoothing is None or len(y_vals) < 2 or any(t <= last_train_year for t in targets):
        return _ols(list(range(last_train_year - len(y_vals) + 1, last_train_year + 1)), y_vals, targets)
    fit = SimpleExpSmoothing(np.asarray(y_vals)).fit(optimized=True)
    max_h = max(t - last_train_year for t in targets)
    fc = fit.forecast(max_h)
    preds = [float(fc[h - 1]) for h in [t - last_train_year for t in targets]]
    return preds, np.array(fit.fittedvalues, dtype=float)

def fmt_int(x):
    if pd.isna(x): return ""
    try: return f"{int(round(float(x))):,}"
    except Exception: return x

@st.cache_data(show_spinner=False)
def load_and_prepare(files, repo_use: bool) -> pd.DataFrame:
    """
    여러 엑셀을 읽어 ‘산업용’만 필터하고 (업종분류, 업종, 연도, 월, 사용량) Long 형태 반환.
    """
    targets: list[tuple[str, BytesIO | Path]] = []
    if files:
        for f in files: targets.append((f.name, f))
    elif repo_use:
        for p in sorted([p for p in Path(".").glob("산업용_*.xlsx")]):
            targets.append((p.name, p))
    if not targets:
        return pd.DataFrame(columns=["업종분류","업종","연도","월","사용량"])

    frames = []
    for name, src in targets:
        df = pd.read_excel(src, engine="openpyxl")
        df.columns = [_clean_col(c) for c in df.columns]
        if not all(c in df.columns for c in ["상품명","업종","판매량"]):
            continue

        col_item, col_ind, col_qty = "상품명", "업종", "판매량"
        col_group = "업종분류" if "업종분류" in df.columns else None
        col_ym = None
        for c in ("판매년월","년월","월","연도","년도"):
            if c in df.columns:
                col_ym = c; break

        item_norm = df[col_item].astype(str).str.replace(r"\s+","", regex=True)
        mask_industry = item_norm == "산업용"
        use_cols = [col_ind, col_qty] + ([col_group] if col_group else []) + ([col_ym] if col_ym else [])
        d = df.loc[mask_industry, use_cols].copy()

        d[col_qty] = _coerce_num(d[col_qty])

        if col_ym:
            yy, mm = _parse_year_month(d[col_ym])
        else:
            yy = pd.Series([pd.NA]*len(d), dtype="Int64")
            mm = pd.Series([pd.NA]*len(d), dtype="Int64")

        if yy.isna().all():
            fn_year = _extract_year_from_filename(name)
            if fn_year is not None:
                yy = pd.Series([fn_year]*len(d), dtype="Int64")
        d["연도"] = yy.astype("Int64")
        d["월"]  = mm.astype("Int64")

        if col_group is None:
            d["업종분류"] = "미지정"
        else:
            d.rename(columns={col_group:"업종분류"}, inplace=True)
        d.rename(columns={col_ind:"업종", col_qty:"사용량"}, inplace=True)

        d = d.dropna(subset=["업종","사용량","연도"])
        d["연도"] = d["연도"].astype(int)
        # 월이 비어있으면 0으로 두지 않고 제거(연간만 있는 경우는 월로 쪼개지 않음)
        frames.append(d[["업종분류","업종","연도","월","사용량"]])

    if not frames:
        return pd.DataFrame(columns=["업종분류","업종","연도","월","사용량"])

    longdf = pd.concat(frames, ignore_index=True)

    # 월 없는 행(월 NaN)은 제외하고 연간만 있는 데이터가 있었다면 그 자체로 남지 않도록 함
    longdf = longdf.dropna(subset=["월"])
    longdf["월"] = longdf["월"].astype(int)

    # 업종분류·업종·연도·월별 합계
    agg = (
        longdf.groupby(["업종분류","업종","연도","월"], as_index=False)["사용량"]
              .sum()
              .sort_values(["연도","월","사용량"], ascending=[True, True, False])
    )
    return agg

# ───────────────────────── 데이터 준비 ─────────────────────────
df_month_all = load_and_prepare(ups, use_repo)

if df_month_all.empty:
    st.info("좌측에서 연도별 엑셀을 올리거나 ‘Repo 파일 자동 읽기’를 켜줘.")
else:
    yr_min, yr_max = df_month_all["연도"].min(), df_month_all["연도"].max()
    st.success(f"로드 완료: 업종 {df_month_all['업종'].nunique():,}개, 업종분류 {df_month_all['업종분류'].nunique():,}종, 연도 범위 {yr_min}–{yr_max} · 지표: 판매량")

# 업종분류 선택(메인영역)
cat_options = ["전체"] + (sorted(df_month_all["업종분류"].dropna().unique().tolist()) if not df_month_all.empty else [])
cat_pick = st.radio("업종분류 선택", cat_options, index=0, horizontal=True, key="cat_pick")

# ───────────────────────── 학습/예측 기간 ─────────────────────────
TRAIN_YEARS = []
FORECAST_YEARS = []
run_clicked = False
if not df_month_all.empty:
    years_all = sorted(df_month_all["연도"].unique().tolist())
    default_2020 = [y for y in years_all if y >= 2020] or years_all
    with st.sidebar:
        st.divider()
        st.header("🗓️ ③ 학습/예측 기간")
        c1, c2 = st.columns(2)
        if c1.button("2020~선택", use_container_width=True):
            st.session_state["train_years"] = default_2020
        if c2.button("전체해제", use_container_width=True):
            st.session_state["train_years"] = []

        TRAIN_YEARS = st.multiselect(
            "학습 연도",
            years_all,
            default=st.session_state.get("train_years", default_2020),
            key="train_years"
        )
        TRAIN_YEARS = sorted(TRAIN_YEARS) if TRAIN_YEARS else []

        # 연간 예측 범위 (기본 2025~2028)
        future = list(range(years_all[-1], years_all[-1] + 10))
        yr_opts = sorted(set(years_all + future))
        start_y = st.selectbox("예측 시작(연)", yr_opts, index=yr_opts.index(years_all[-1]))
        end_y   = st.selectbox("예측 종료(연)", yr_opts, index=min(len(yr_opts)-1, yr_opts.index(years_all[-1])+3))
        FORECAST_YEARS = list(range(min(start_y, end_y), max(start_y, end_y) + 1))

        st.caption("※ 2025년은 9–12월을 월별 추세로 보정한 **연간(추정)** 값으로 사용.")
        st.divider()
        run_clicked = st.button("🚀 예측 시작", use_container_width=True)

# ───────────────────────── 상태/계산 ─────────────────────────
if "started" not in st.session_state: st.session_state.started = False
if "store"   not in st.session_state: st.session_state.store   = {}

def compute_and_store():
    # 1) 업종분류 필터
    if cat_pick == "전체":
        dfm = df_month_all.copy()
    else:
        dfm = df_month_all.loc[df_month_all["업종분류"] == cat_pick].copy()

    if dfm.empty:
        st.warning("선택한 업종분류에 데이터가 없습니다.")
        st.stop()

    # 2) 2025의 마지막 실측 월
    last_yr = max(dfm["연도"])
    max_month_2025 = int(dfm.loc[dfm["연도"] == 2025, "월"].max()) if (dfm["연도"] == 2025).any() else 0
    missing_months_2025 = [m for m in range(max_month_2025+1, 13) if m >= 1 and m <= 12]

    # 3) 업종×월×연도 피벗(월별)
    pm = dfm.pivot_table(index=["업종","연도"], columns="월", values="사용량", aggfunc="sum").fillna(0)

    # 4) 2025 연간(추정) 만들기: 실측(1..max_m) + 예측(>max_m)
    est25 = {}
    for ind in pm.index.get_level_values(0).unique():
        # 실측 YTD 합
        ytd = 0.0
        if 2025 in pm.loc[ind].index:
            ytd = float(pm.loc[(ind, 2025)].loc[range(1, max_month_2025+1)].sum()) if max_month_2025>0 else 0.0

        # 누락 월 예측
        miss_sum = 0.0
        for m in missing_months_2025:
            # 해당 월의 학습 연도들에서 값 모으기 (같은 달만)
            ys = []
            xs = []
            for y in TRAIN_YEARS:
                if (ind, y) in pm.index:
                    val = pm.loc[(ind, y)].get(m, np.nan)
                    if pd.notna(val):
                        xs.append(y); ys.append(float(val))
            # 2025에 해당 월이 이미 있으면(이론상 없음) 포함
            if (ind, 2025) in pm.index and m <= max_month_2025:
                xs.append(2025); ys.append(float(pm.loc[(ind, 2025)].get(m, 0.0)))

            if len(xs) >= 2:
                xs_sorted, ys_sorted = zip(*sorted(zip(xs, ys)))
                preds, _ = _holt_level(list(ys_sorted), list(xs_sorted), [2025], damped=True)
                miss_sum += max(0.0, preds[0])  # 음수 방지용 max
            elif len(xs) == 1:
                miss_sum += max(0.0, ys[0])
            else:
                miss_sum += 0.0

        est25[ind] = ytd + miss_sum

    est25_s = pd.Series(est25, name=2025)

    # 5) 연간 피벗(업종×연도) — 2025연간을 추정치로 대체
    pa = dfm.groupby(["업종","연도"], as_index=False)["사용량"].sum().pivot(index="업종", columns="연도", values="사용량").fillna(0)
    if 2025 in pa.columns:
        pa[2025] = pa.index.map(est25_s).fillna(pa[2025]).astype(float)
    else:
        pa[2025] = pa.index.map(est25_s).fillna(0.0).astype(float)

    # 6) 학습 연도 확인
    missing = [y for y in TRAIN_YEARS if y not in pa.columns]
    if missing:
        st.error(f"학습 연도 데이터 없음: {missing}")
        st.stop()

    pv = pa.reindex(columns=sorted(set(TRAIN_YEARS))).fillna(0)  # 학습 테이블

    # 7) 예측 결과 테이블(연간)
    result = pv.copy()
    # 연도 실적 라벨링: 2025는 (추정) 표기
    renamed_cols = []
    for c in result.columns:
        if int(c) == 2025:
            renamed_cols.append(f"{c} 실적(추정)")
        else:
            renamed_cols.append(f"{c} 실적")
    result.columns = renamed_cols

    # 예측(연간)
    for ind, row in pv.iterrows():
        y = row.values.astype(float).tolist()
        x = pv.columns.astype(int).tolist()
        last = x[-1]
        for m in methods:
            label = str(m)
            if "OLS" in label:
                preds, _ = _ols(x, y, FORECAST_YEARS)
            elif "CAGR" in label:
                preds, _ = _cagr(x, y, FORECAST_YEARS)
            elif "Holt" in label:
                preds, _ = _holt_annual(y, last, FORECAST_YEARS, damped=False)
            elif "SES" in label:
                preds, _ = _ses(y, last, FORECAST_YEARS)
            else:
                preds, _ = _ols(x, y, FORECAST_YEARS)
            for yy, p in zip(FORECAST_YEARS, preds):
                col = f"{label}({yy})"
                if col not in result.columns: result[col] = np.nan
                result.loc[ind, col] = p

    sort_method = methods[0]
    sort_col = f"{sort_method}({FORECAST_YEARS[-1]})"
    if sort_col not in result.columns:
        alt_cols = [c for c in result.columns if c.endswith(f"({FORECAST_YEARS[-1]})")]
        sort_col = alt_cols[0] if alt_cols else result.columns[-1]
    final_sorted = result.sort_values(by=sort_col, ascending=False)
    final_sorted.index.name = "업종"

    totals = final_sorted.sum(axis=0, numeric_only=True)
    total_row = pd.DataFrame([totals.to_dict()], index=["총합"])
    final_with_total = pd.concat([final_sorted, total_row], axis=0)

    st.session_state.store = dict(
        pv=pv, final=final_sorted, final_total=final_with_total,
        train_years=sorted(set(TRAIN_YEARS)), fc_years=FORECAST_YEARS, methods=methods,
        cat_pick=cat_pick
    )
    st.session_state.started = True

if run_clicked and not df_month_all.empty and methods and TRAIN_YEARS and FORECAST_YEARS:
    compute_and_store()

# ───────────────────────── 결과 표시 ─────────────────────────
if st.session_state.started:
    pv = st.session_state.store["pv"]
    final_sorted = st.session_state.store["final"]
    final_total  = st.session_state.store["final_total"]
    TRAIN_YEARS  = st.session_state.store["train_years"]
    FORECAST_YEARS = st.session_state.store["fc_years"]
    methods = st.session_state.store["methods"]
    cat_pick = st.session_state.store["cat_pick"]

    cat_text = "전체" if cat_pick == "전체" else f"{cat_pick}"
    st.success(f"[업종분류: {cat_text}] 업종 {pv.shape[0]}개, 학습 {min(TRAIN_YEARS)}–{max(TRAIN_YEARS)}, 예측 {FORECAST_YEARS[0]}–{FORECAST_YEARS[-1]} (※ 2025은 9–12월 추정 포함)")

    # 표
    st.subheader("🧾 업종별 예측 표")
    disp = final_total.copy()
    disp.insert(0, "업종", disp.index)
    for c in disp.columns[1:]: disp[c] = disp[c].apply(fmt_int)
    st.dataframe(disp.reset_index(drop=True), use_container_width=True)

    # 총합 그래프 (실적+예측 포인트)
    st.subheader("📈 연도별 총합(실적 라인 + 예측 포인트)")
    tot_actual = pv.sum(axis=0).reset_index()
    tot_actual.columns = ["연도", "합계"]
    pts = []
    for m in methods:
        for yy in FORECAST_YEARS:
            col = f"{m}({yy})"
            if col in final_sorted.columns:
                pts.append({"연도": yy, "방법": m, "값": float(final_sorted[col].sum())})
    pts_df = pd.DataFrame(pts)

    area = alt.Chart(tot_actual).mark_area(opacity=0.25).encode(
        x=alt.X("연도:O", title="연도"),
        y=alt.Y("합계:Q", title="총합(실적·2025은 추정포함)", axis=alt.Axis(format=",")),
        tooltip=[alt.Tooltip("연도:O"), alt.Tooltip("합계:Q", format=",")]
    )
    line = alt.Chart(tot_actual).mark_line(size=3).encode(x="연도:O", y=alt.Y("합계:Q", axis=alt.Axis(format=",")))
    if not pts_df.empty:
        ptsch = alt.Chart(pts_df).mark_point(size=150, filled=True).encode(
            x="연도:O", y=alt.Y("값:Q", axis=alt.Axis(format=",")),
            color=alt.Color("방법:N", legend=alt.Legend(title="방법")),
            shape=alt.Shape("방법:N"),
            tooltip=[alt.Tooltip("방법:N"), alt.Tooltip("연도:O"), alt.Tooltip("값:Q", format=",")]
        )
        labels = ptsch.mark_text(dy=-12, fontWeight="bold").encode(text=alt.Text("값:Q", format=","))
        st.altair_chart((area + line + ptsch + labels).interactive(), use_container_width=True, theme="streamlit")
    else:
        st.altair_chart((area + line).interactive(), use_container_width=True, theme="streamlit")

    # Top-10 — 막대/라인
    st.subheader("🏆 상위 10개 업종 — 예측 비교 / 실적 추이")
    yy_pick = st.radio("막대그래프 기준 예측연도", FORECAST_YEARS, index=len(FORECAST_YEARS)-1, horizontal=True, key="yy_pick")

    method_cols = [f"{m}({yy_pick})" for m in methods if f"{m}({yy_pick})" in final_sorted.columns]
    if method_cols:
        top10 = final_sorted.head(10).index.tolist()
        bar_base = final_sorted.loc[top10, method_cols].copy()
        bar_base.index.name = "업종"
        pred_long = bar_base.reset_index().melt(id_vars="업종", var_name="방법연도", value_name="예측")
        pred_long["방법"] = pred_long["방법연도"].str.replace(r"\(\d{4}\)$", "", regex=True)

        sel = alt.selection_point(fields=["방법"], bind="legend")
        bars = alt.Chart(pred_long).mark_bar().encode(
            x=alt.X("업종:N", sort=top10, title=None),
            xOffset=alt.XOffset("방법:N"),
            y=alt.Y("예측:Q", axis=alt.Axis(format=","), title=f"{yy_pick} 예측"),
            color=alt.Color("방법:N", legend=alt.Legend(title="방법")),
            opacity=alt.condition(sel, alt.value(1.0), alt.value(0.25)),
            tooltip=[alt.Tooltip("업종:N"), alt.Tooltip("방법:N"), alt.Tooltip("예측:Q", format=",")]
        ).add_params(sel).properties(height=420)
        bar_txt = bars.mark_text(dy=-5, fontSize=10).encode(text=alt.Text("예측:Q", format=","))

        st.markdown("※ 라인 그래프는 **첫 번째로 선택한 방법**으로 예측 연도를 이어서 보여줘.")
        method_for_line = methods[0]
        pred_cols_for_line = [f"{method_for_line}({yy})" for yy in FORECAST_YEARS if f"{method_for_line}({yy})" in final_sorted.columns]

        actual_long = pv.loc[top10, TRAIN_YEARS].reset_index().melt(id_vars="업종", var_name="연도", value_name="값")
        actual_long["출처"] = "실적"

        pred_line = pd.DataFrame()
        if pred_cols_for_line:
            import re as _re
            pred_line = final_sorted.loc[top10, pred_cols_for_line].copy()
            pred_line.columns = [int(_re.search(r"\((\d{4})\)", c).group(1)) for c in pred_line.columns]
            pred_line = pred_line.reset_index().melt(id_vars="업종", var_name="연도", value_name="값")
            pred_line["출처"] = f"예측({method_for_line})"

        line_df = pd.concat([actual_long, pred_line], ignore_index=True)
        year_order = TRAIN_YEARS + [y for y in FORECAST_YEARS if y not in TRAIN_YEARS]
        line_df["연도"] = line_df["연도"].astype(str)
        line_df["연도"] = pd.Categorical(line_df["연도"], categories=[str(y) for y in year_order], ordered=True)

        sel2 = alt.selection_point(fields=["업종"], bind="legend")
        lines = alt.Chart(line_df).mark_line(point=True, strokeWidth=3).encode(
            x=alt.X("연도:O", title=None),
            y=alt.Y("값:Q", axis=alt.Axis(format=",")),
            color=alt.Color("업종:N", sort=top10, legend=alt.Legend(title="업종(클릭으로 강조)")),
            opacity=alt.condition(sel2, alt.value(1.0), alt.value(0.25)),
            tooltip=[alt.Tooltip("업종:N"), alt.Tooltip("연도:O"), alt.Tooltip("값:Q", format=","), alt.Tooltip("출처:N")]
        ).add_params(sel2).properties(height=420)

        c1, c2 = st.columns(2)
        with c1: st.altair_chart((bars + bar_txt).interactive(), use_container_width=True, theme="streamlit")
        with c2: st.altair_chart(lines.interactive(), use_container_width=True, theme="streamlit")
    else:
        st.info(f"{yy_pick}년 예측 열이 없어서 막대그래프는 건너뛰었어.")

    # 다운로드
    st.subheader("💾 다운로드")
    out_all = final_total.copy()
    out_all.insert(0, "업종", out_all.index)
    fname = f"industry_forecast_{FORECAST_YEARS[0]}-{FORECAST_YEARS[-1]}.xlsx"

    wb = Workbook(); wb.remove(wb.active)
    ws_all = wb.create_sheet("전체")
    for r in dataframe_to_rows(out_all, index=False, header=True): ws_all.append(r)

    def add_method_sheet(mth):
        ws = wb.create_sheet(mth)
        dfm = pv.copy(); dfm.columns = [f"{c} 실적" if "2025" not in str(c) else f"{c} 실적(추정)" for c in dfm.columns]
        pred_cols = [f"{mth}({yy})" for yy in FORECAST_YEARS if f"{mth}({yy})" in final_sorted.columns]
        for c in pred_cols: dfm[c] = final_sorted[c]
        order_col = pred_cols[-1] if pred_cols else dfm.columns[-1]
        dfm = dfm.sort_values(by=order_col, ascending=False).reset_index().rename(columns={"index":"업종"})
        for r in dataframe_to_rows(dfm, index=False, header=True): ws.append(r)

        # Top-20 막대(종료연도)
        if pred_cols:
            topN = min(20, len(dfm))
            y = FORECAST_YEARS[-1]; use_col = f"{mth}({y})"
            sc = dfm.shape[1] + 2
            ws.cell(row=1, column=sc, value="업종")
            ws.cell(row=1, column=sc+1, value=f"{y} 예측")
            for i in range(topN):
                ws.cell(row=i+2, column=sc, value=dfm.loc[i, "업종"])
                v = dfm.loc[i, use_col]
                ws.cell(row=i+2, column=sc+1, value=float(v if pd.notna(v) else 0))
            bar = BarChart(); bar.title = f"Top-20 {y} ({mth})"
            data = Reference(ws, min_col=sc+1, min_row=1, max_row=topN+1)
            cats = Reference(ws, min_col=sc,   min_row=2, max_row=topN+1)
            bar.add_data(data, titles_from_data=True); bar.set_categories(cats)
            bar.y_axis.number_format = '#,##0'
            ws.add_chart(bar, ws.cell(row=2, column=sc+3).coordinate)

        # 연도별 총합 라인
        la = dfm.shape[1] + 8
        ws.cell(row=1, column=la,   value="연도")
        ws.cell(row=1, column=la+1, value="총합")
        for i, y in enumerate(TRAIN_YEARS, start=2):
            ws.cell(row=i, column=la,   value=y)
            ws.cell(row=i, column=la+1, value=float(pv[y].sum()))
        base = len(TRAIN_YEARS) + 2
        for j, y in enumerate(FORECAST_YEARS):
            ws.cell(row=base+j, column=la,   value=y)
            col = f"{mth}({y})"
            tot = float(final_sorted[col].sum()) if col in final_sorted.columns else 0.0
            ws.cell(row=base+j, column=la+1, value=tot)
        lch = LineChart(); lch.title = f"연도별 총합(실적+예측, {mth})"
        mr = base + len(FORECAST_YEARS) - 1
        d = Reference(ws, min_col=la+1, min_row=1, max_row=mr)
        c = Reference(ws, min_col=la,   min_row=2, max_row=mr)
        lch.add_data(d, titles_from_data=True); lch.set_categories(c)
        lch.y_axis.number_format = '#,##0'
        ws.add_chart(lch, ws.cell(row=2, column=la+3).coordinate)

    for m in methods: add_method_sheet(m)

    bio = BytesIO(); wb.save(bio)
    st.download_button("엑셀(xlsx) 다운로드", bio.getvalue(), file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    st.download_button(
        "업종별 예측표 CSV 다운로드",
        out_all.to_csv(index=False).encode("utf-8-sig"),
        file_name=fname.replace(".xlsx",".csv"),
        mime="text/csv"
    )
